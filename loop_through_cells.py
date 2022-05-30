import openpyxl as xl

# open an existing workbook
wb = xl.load_workbook(filename="maven_ski_shop_data.xlsx")

# set a ws
ws = wb["Inventory_Levels"]

# loop through cells
for i, cell in enumerate(ws["B"], start=1):
    if i == 1:
        ws[f"C{i}"] = "Inventory Status"
    elif cell.value > 5:
        ws[f"C{i}"] = "Healthy Stock"
    elif cell.value > 0:
        ws[f"C{i}"] = "Low Stock"
    else:
        ws[f"C{i}"] = "Out of Stock"

# save to another wb
wb.save("maven_ski_shop_inventory_update.xlsx")
